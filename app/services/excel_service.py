"""
Excel export service — generates VAT summary reports.
Output format mirrors BẢNG KÊ HÓA ĐƠN CHỨNG TỪ HÀNG HÓA, DỊCH VỤ MUA VÀO / BÁN RA.

Fixes:
  - Sheet title set correctly (MUA VÀO / BÁN RA)
  - mat_hang resolved from line_items_json when empty
  - Company header (name / tax code / address) read from AppSetting
  - _safe_sheet() strips invalid Excel sheet-name characters

New:
  - export_by_year(): gộp các batch (tháng) đã chốt trong một năm cho một
    tài khoản vào một file Excel duy nhất, mỗi tháng/batch là một nhóm có
    hàng TỔNG cắt ngang — ở cả sheet MUA VÀO và BÁN RA.
"""
from __future__ import annotations

import json
import re
from collections import defaultdict
from datetime import datetime
from pathlib import Path
from typing import List, Optional

from loguru import logger
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from app.config import Config
from app.db.models import Invoice
from app.db.repository import InvoiceRepository, SettingsRepository

# ── Style constants ────────────────────────────────────────────────────────
_THIN   = Side(style="thin")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

_BLUE_FILL  = PatternFill("solid", start_color="1F4E79", end_color="1F4E79")
_TOTAL_FILL = PatternFill("solid", start_color="FFEB9C", end_color="FFEB9C")
_ALT_FILL   = PatternFill("solid", start_color="F2F2F2", end_color="F2F2F2")

_WHITE_BOLD   = Font(name="Arial", bold=True, color="FFFFFF", size=10)
_BLACK_NORM   = Font(name="Arial", size=10)
_TITLE_FONT   = Font(name="Arial", bold=True, size=13)
_COMPANY_BOLD = Font(name="Arial", bold=True, size=10)

_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
_LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)
_RIGHT  = Alignment(horizontal="right",  vertical="center")

# ── Column definitions ─────────────────────────────────────────────────────
_MUA_VAO_COLS = [
    ("STT",                              6,  "stt"),
    ("Ký hiệu mẫu số",                 10,  "invoice_form"),
    ("Ký hiệu hóa đơn",               14,  "invoice_symbol"),
    ("Số hóa đơn",                     14,  "invoice_no"),
    ("Ngày lập",                       13,  "issue_date_fmt"),
    ("Tên người bán",                  36,  "seller_name"),
    ("MST người bán",                  18,  "seller_tax_code"),
    ("Mặt hàng",                       36,  "mat_hang_resolved"),
    ("Doanh số mua vào\nchưa có thuế", 18,  "amount"),
    ("Thuế suất %",                    10,  "vat_rate"),
    ("Thuế GTGT",                      15,  "vat_amount"),
    ("Thành tiền VND",                 16,  "total_amount"),
    ("Tháng kê khai",                  14,  "thang_ke_khai"),
    ("Chứng từ thanh toán\n(TM/CK)",   22,  "payment_note"),
    ("Ngân hàng",                      12,  "bank_name"),
    ("Điều chỉnh",                     12,  "adj"),
    ("Bị điều chỉnh",                  12,  "adj2"),
    ("Ghi chú",                        22,  "ghi_chu"),
    ("Mã công trình",                  16,  "ma_cong_trinh"),
    ("Số hợp đồng",                    24,  "so_hop_dong"),
    ("Ngày hợp đồng",                  14,  "ngay_hop_dong"),
    ("HĐ bán ra\ntương ứng",           20,  "hd_ban_ra_tuong_ung"),
]

_BAN_RA_COLS = [
    ("STT",                             6,  "stt"),
    ("Ký hiệu mẫu số",                10,  "invoice_form"),
    ("Ký hiệu hóa đơn",              14,  "invoice_symbol"),
    ("Số HĐ",                         14,  "invoice_no"),
    ("Ngày HĐ",                       13,  "issue_date_fmt"),
    ("Tên người mua",                 36,  "buyer_name"),
    ("Mã số thuế",                    18,  "buyer_tax_code"),
    ("Mặt hàng",                      36,  "mat_hang_resolved"),
    ("Doanh số bán\nchưa thuế",       18,  "amount"),
    ("Thuế suất",                     10,  "vat_rate"),
    ("Thuế GTGT",                     15,  "vat_amount"),
    ("Tổng thanh toán",               16,  "total_amount"),
    ("Điều chỉnh",                    12,  "adj"),
    ("Bị điều chỉnh",                 12,  "adj2"),
    ("Mã công trình",                 16,  "ma_cong_trinh"),
    ("Số hợp đồng",                   24,  "so_hop_dong"),
    ("Ngày hợp đồng",                 14,  "ngay_hop_dong"),
    ("Chứng từ thanh toán\n(TM/CK)",  22,  "payment_note"),
    ("Ngân hàng",                     12,  "bank_name"),
    ("Tháng kê khai",                 14,  "thang_ke_khai"),
    ("Ghi chú",                       22,  "ghi_chu"),
]


# ── Batch label parsing (dùng cho export theo năm) ─────────────────────────
# Khớp với format được sinh ở snapshot.html: `Tháng ${month}/${year}`
_BATCH_MONTH_RE = re.compile(r"(\d{1,2})\s*/\s*(\d{4})")


def _parse_batch_month_year(label: str):
    """Trích (tháng, năm) từ batch label dạng 'Tháng 3/2026'. Trả (None, None) nếu không khớp."""
    if not label:
        return None, None
    m = _BATCH_MONTH_RE.search(label)
    if not m:
        return None, None
    return int(m.group(1)), int(m.group(2))


# ── Company info ───────────────────────────────────────────────────────────

def _get_company_info(account_id: Optional[int] = None) -> dict:
    """
    Đọc thông tin công ty theo thứ tự ưu tiên:
      1. GdtAccount.company_* (nếu account_id được truyền và có dữ liệu)
      2. AppSetting global (company_name, company_tax_code, company_address, company_report_title)
      3. Rỗng
    """
    acct_info = {}
    if account_id is not None:
        try:
            from app.db.repository import GdtAccountRepository
            acct = GdtAccountRepository.get_by_id(account_id)
            if acct:
                acct_info = {
                    "name":         acct.company_name         or "",
                    "tax_code":     acct.company_tax_code     or "",
                    "address":      acct.company_address      or "",
                    "report_title": acct.company_report_title or "",
                }
        except Exception:
            pass

    try:
        global_info = {
            "name":         SettingsRepository.get("company_name",         "") or "",
            "tax_code":     SettingsRepository.get("company_tax_code",     "") or "",
            "address":      SettingsRepository.get("company_address",      "") or "",
            "report_title": SettingsRepository.get("company_report_title", "") or "",
        }
    except Exception:
        global_info = {"name": "", "tax_code": "", "address": "", "report_title": ""}

    # Merge: account takes priority over global for each field
    return {
        "name":         acct_info.get("name")         or global_info["name"],
        "tax_code":     acct_info.get("tax_code")     or global_info["tax_code"],
        "address":      acct_info.get("address")      or global_info["address"],
        "report_title": acct_info.get("report_title") or global_info["report_title"],
    }


# ── Main service ───────────────────────────────────────────────────────────

class ExcelService:

    @staticmethod
    def export_vat_summary(
        start_date: Optional[str] = None,
        end_date: Optional[str] = None,
        ghi_chu: Optional[str] = None,
        thang_ke_khai: Optional[str] = None,
        search: Optional[str] = None,
        invoice_category: Optional[str] = None,
        account_id: Optional[int] = None,
    ) -> Path:
        Config.EXPORT_PATH.mkdir(parents=True, exist_ok=True)

        company = _get_company_info(account_id=account_id)
        if not company["name"] and not company["tax_code"]:
            raise ValueError("Vui lòng chọn công ty trước khi xuất Excel.")

        invoices = InvoiceRepository.get_for_export(
            start_date=_parse_date(start_date),
            end_date=_parse_date(end_date),
            ghi_chu=ghi_chu,
            thang_ke_khai=thang_ke_khai,
            search=search,
            account_id=account_id,
        )

        mua_vao = [i for i in invoices if not (i.invoice_category or "").startswith("sale_")]
        ban_ra  = [i for i in invoices if     (i.invoice_category or "").startswith("sale_")]

        company   = _get_company_info(account_id=account_id)
        dest      = Config.EXPORT_PATH / f"bang_ke_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

        wb = Workbook()
        ws_mua       = wb.active
        ws_mua.title = "MUA VÀO"
        _build_sheet(ws_mua, "MUA VÀO", mua_vao, _MUA_VAO_COLS, is_purchase=True,  company=company)

        ws_ban       = wb.create_sheet("BÁN RA")
        _build_sheet(ws_ban, "BÁN RA",  ban_ra,  _BAN_RA_COLS,  is_purchase=False, company=company)

        ws_vat       = wb.create_sheet("VAT")
        _build_vat_sheet(ws_vat, invoices, company)

        wb.save(str(dest))
        logger.info("Excel report saved: {}", dest)
        return dest

    @staticmethod
    def export_from_list(
        invoices: List[Invoice],
        label: str = "Hóa đơn",
        account_id: Optional[int] = None,
    ) -> Path:
        Config.EXPORT_PATH.mkdir(parents=True, exist_ok=True)

        mua_vao = [i for i in invoices if not (i.invoice_category or "").startswith("sale_")]
        ban_ra  = [i for i in invoices if     (i.invoice_category or "").startswith("sale_")]

        company = _get_company_info(account_id=account_id)
        if not company["name"] and not company["tax_code"]:
            raise ValueError("Vui lòng chọn công ty trước khi xuất Excel.")

        safe      = _safe_sheet(label)
        dest      = Config.EXPORT_PATH / f"snapshot_{_safe_filename(label)}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

        wb = Workbook()
        ws_mua       = wb.active
        ws_mua.title = f"MUA VÀO - {safe}"[:31]
        _build_sheet(ws_mua, f"MUA VÀO — {label}", mua_vao, _MUA_VAO_COLS, is_purchase=True,  company=company)

        ws_ban       = wb.create_sheet(f"BÁN RA - {safe}"[:31])
        _build_sheet(ws_ban, f"BÁN RA — {label}",  ban_ra,  _BAN_RA_COLS,  is_purchase=False, company=company)

        ws_vat       = wb.create_sheet("VAT")
        _build_vat_sheet(ws_vat, invoices, company)

        wb.save(str(dest))
        logger.info("Snapshot Excel saved: {}", dest)
        return dest

    @staticmethod
    def export_by_year(year: int, account_id: int) -> Path:
        """
        Xuất toàn bộ hóa đơn đã chốt theo các batch (tháng) trong một năm,
        cho một tài khoản. Mỗi tháng/batch là một nhóm, có hàng TỔNG cắt
        ngang giữa các tháng — ở cả sheet MUA VÀO và BÁN RA.

        Batch label phải có dạng "Tháng M/YYYY" (đúng format được sinh tự
        động khi chốt batch trong UI) thì mới được nhận diện thuộc năm nào.
        """
        from app.db.repository import SnapshotRepository  # tránh import vòng

        Config.EXPORT_PATH.mkdir(parents=True, exist_ok=True)

        company = _get_company_info(account_id=account_id)
        if not company["name"] and not company["tax_code"]:
            raise ValueError("Vui lòng chọn công ty trước khi xuất Excel.")

        batches = SnapshotRepository.get_all(account_id=account_id)

        year_batches = []
        for b in batches:
            m, y = _parse_batch_month_year(b.label)
            if y == year:
                year_batches.append((m or 0, b))
        year_batches.sort(key=lambda t: t[0])

        if not year_batches:
            raise ValueError(f"Không có batch nào trong năm {year} cho công ty đã chọn.")

        all_invoices: List[Invoice] = []
        mua_groups: list = []
        ban_groups: list = []

        for _, b in year_batches:
            inv_list = SnapshotRepository.get_batch_for_export(b.id)
            if not inv_list:
                continue
            all_invoices.extend(inv_list)
            mua = [i for i in inv_list if not (i.invoice_category or "").startswith("sale_")]
            ban = [i for i in inv_list if     (i.invoice_category or "").startswith("sale_")]
            if mua:
                mua_groups.append((b.label, mua))
            if ban:
                ban_groups.append((b.label, ban))

        if not all_invoices:
            raise ValueError(f"Không có hóa đơn nào trong năm {year}.")

        dest = Config.EXPORT_PATH / f"bang_ke_nam_{year}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

        wb = Workbook()
        ws_mua       = wb.active
        ws_mua.title = "MUA VÀO"
        mua_flat     = [i for _, lst in mua_groups for i in lst]
        _build_sheet(ws_mua, f"MUA VÀO — Năm {year}", mua_flat, _MUA_VAO_COLS,
                     is_purchase=True, company=company, groups=mua_groups)

        ws_ban   = wb.create_sheet("BÁN RA")
        ban_flat = [i for _, lst in ban_groups for i in lst]
        _build_sheet(ws_ban, f"BÁN RA — Năm {year}", ban_flat, _BAN_RA_COLS,
                     is_purchase=False, company=company, groups=ban_groups)

        ws_vat = wb.create_sheet("VAT")
        _build_vat_sheet(ws_vat, all_invoices, company)

        wb.save(str(dest))
        logger.info("Yearly Excel report saved: {} ({} batches)", dest, len(year_batches))
        return dest


# ── Sheet builder ──────────────────────────────────────────────────────────

def _build_sheet(ws, sheet_label: str, invoices: List[Invoice],
                 col_defs: list, is_purchase: bool, company: dict,
                 groups: Optional[list] = None) -> None:
    """
    groups: list các tuple (label, invoices) đã được nhóm sẵn từ bên ngoài
    (vd. theo batch/tháng cho export theo năm). Nếu None, tự nhóm theo
    `thang_ke_khai` như hành vi cũ (_group_by_ky).
    """
    num_cols = len(col_defs)
    last_col = get_column_letter(num_cols)

    def _merge_row(row: int, value: str, font, height: int = 15):
        ws.merge_cells(f"A{row}:{last_col}{row}")
        c = ws[f"A{row}"]
        c.value     = value
        c.font      = font
        c.alignment = _LEFT
        ws.row_dimensions[row].height = height

    # Rows 1-3: company info
    _merge_row(1, f"Tên công ty : {company['name']}",     _COMPANY_BOLD)
    _merge_row(2, f"Mã số thuế  : {company['tax_code']}", _BLACK_NORM)
    _merge_row(3, f"Địa chỉ     : {company['address']}",  _BLACK_NORM)

    # Row 4: report title
    ws.merge_cells(f"A4:{last_col}4")
    c4           = ws["A4"]
    c4.value     = "BẢNG KÊ HÓA ĐƠN CHỨNG TỪ HÀNG HÓA, DỊCH VỤ " + ("MUA VÀO" if is_purchase else "BÁN RA")
    c4.font      = _TITLE_FONT
    c4.alignment = _CENTER
    ws.row_dimensions[4].height = 26

    # Row 5: timestamp
    ws.merge_cells(f"A5:{last_col}5")
    c5           = ws["A5"]
    c5.value     = f"Xuất ngày: {datetime.now().strftime('%d/%m/%Y %H:%M')}   —   Tổng: {len(invoices)} hóa đơn"
    c5.font      = _BLACK_NORM
    c5.alignment = _CENTER
    ws.row_dimensions[5].height = 14

    # Row 6: column headers
    for col_idx, (hdr, width, _) in enumerate(col_defs, 1):
        cell           = ws.cell(row=6, column=col_idx, value=hdr)
        cell.font      = _WHITE_BOLD
        cell.fill      = _BLUE_FILL
        cell.border    = _BORDER
        cell.alignment = _CENTER
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[6].height = 36

    ws.freeze_panes = "A7"

    # Data rows
    groups = groups if groups is not None else _group_by_ky(invoices)
    current_row  = 7
    grand_amount = grand_vat = grand_total = 0.0

    for ky_label, ky_invoices in groups:
        stt = 1
        ky_amount = ky_vat = ky_total = 0.0

        for inv in ky_invoices:
            row_data = _invoice_to_row(inv, stt, col_defs)
            alt = stt % 2 == 0
            for col_idx, val in enumerate(row_data, 1):
                cell        = ws.cell(row=current_row, column=col_idx, value=val)
                cell.border = _BORDER
                cell.font   = _BLACK_NORM
                if alt:
                    cell.fill = _ALT_FILL
                _, _, field = col_defs[col_idx - 1]
                if field in ("amount", "vat_amount", "total_amount"):
                    cell.alignment     = _RIGHT
                    cell.number_format = '#,##0'
                elif field == "stt":
                    cell.alignment = _CENTER
                else:
                    cell.alignment = _LEFT

            ky_amount += inv.amount       or 0
            ky_vat    += inv.vat_amount   or 0
            ky_total  += inv.total_amount or 0
            stt         += 1
            current_row += 1

        if ky_label:
            _write_subtotal_row(ws, current_row, num_cols, ky_label,
                                ky_amount, ky_vat, ky_total, col_defs)
            current_row += 1

        grand_amount += ky_amount
        grand_vat    += ky_vat
        grand_total  += ky_total

    _write_subtotal_row(ws, current_row, num_cols, "TỔNG CỘNG",
                        grand_amount, grand_vat, grand_total, col_defs, is_grand=True)


def _write_subtotal_row(ws, row, num_cols, label, amount, vat, total,
                        col_defs, is_grand=False):
    fill = _TOTAL_FILL
    font = Font(name="Arial", bold=True, size=10)

    amount_col = vat_col = total_col = None
    for i, (_, _, field) in enumerate(col_defs, 1):
        if field == "amount":       amount_col = i
        if field == "vat_amount":   vat_col    = i
        if field == "total_amount": total_col  = i

    span_end = (amount_col - 1) if amount_col else num_cols
    if span_end >= 1:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span_end)

    lc           = ws.cell(row=row, column=1, value=label if is_grand else f"TỔNG {label}")
    lc.font      = font
    lc.fill      = fill
    lc.border    = _BORDER
    lc.alignment = Alignment(horizontal="right", vertical="center")

    for col_idx in range(1, num_cols + 1):
        c = ws.cell(row=row, column=col_idx)
        c.fill = fill; c.border = _BORDER; c.font = font

    for col_idx, val in [(amount_col, amount), (vat_col, vat), (total_col, total)]:
        if col_idx:
            c = ws.cell(row=row, column=col_idx, value=val)
            c.number_format = '#,##0'
            c.fill = fill; c.border = _BORDER; c.font = font; c.alignment = _RIGHT

    ws.row_dimensions[row].height = 18


def _invoice_to_row(inv: Invoice, stt: int, col_defs: list) -> list:
    date_str = inv.issue_date.strftime("%d/%m/%Y") if inv.issue_date else ""

    mat_hang = ""
    if inv.line_items_json:
        try:
            items = json.loads(inv.line_items_json)
            names = []
            for it in items:
                ten = it.get("ten") or it.get("ten_hhdvu")
                if ten:
                    names.append(ten)
            mat_hang = "; ".join(names)
        except Exception:
            pass
    if not mat_hang:
        mat_hang = inv.mat_hang or ""

    row = []
    for _, _, field in col_defs:
        if   field == "stt":                row.append(stt)
        elif field == "issue_date_fmt":     row.append(date_str)
        elif field == "mat_hang_resolved":  row.append(mat_hang)
        elif field == "amount":             row.append(inv.amount       or 0)
        elif field == "vat_amount":         row.append(inv.vat_amount   or 0)
        elif field == "total_amount":       row.append(inv.total_amount or 0)
        elif field in ("adj", "adj2"):      row.append("")
        else:
            row.append(getattr(inv, field, "") or "")
    return row


def _group_by_ky(invoices: List[Invoice]):
    ordered_keys: list = []
    groups: dict = defaultdict(list)
    for inv in invoices:
        key = inv.thang_ke_khai or ""
        if key not in ordered_keys:
            ordered_keys.append(key)
        groups[key].append(inv)
    return [(k, groups[k]) for k in ordered_keys]


def _safe_sheet(name: str) -> str:
    for ch in r'/\\?*[]:{':
        name = name.replace(ch, "-")
    return name[:31]


def _safe_filename(name: str) -> str:
    for ch in r'/\\?*[]:{} ':
        name = name.replace(ch, "_")
    return name[:40]


def _build_vat_sheet(ws, invoices: List[Invoice], company: dict) -> None:
    """Build the VAT summary sheet (BẢNG KÊ TỔNG HỢP) mirroring the sample template."""
    _HEADER_FILL  = PatternFill("solid", start_color="1F4E79", end_color="1F4E79")
    _MONTH_FILL   = PatternFill("solid", start_color="E2EFDA", end_color="E2EFDA")
    _TOTAL_F      = PatternFill("solid", start_color="FFEB9C", end_color="FFEB9C")
    _CARRY_FILL   = PatternFill("solid", start_color="FCE4D6", end_color="FCE4D6")
    _WHITE_BOLD   = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    _BOLD10       = Font(name="Arial", bold=True, size=10)
    _NORM10       = Font(name="Arial", size=10)
    _TITLE_F      = Font(name="Arial", bold=True, size=13)
    _RIGHT        = Alignment(horizontal="right",  vertical="center")
    _CENTER       = Alignment(horizontal="center", vertical="center", wrap_text=True)
    _LEFT         = Alignment(horizontal="left",   vertical="center", wrap_text=True)
    _NUM_FMT      = '#,##0'

    # Column widths: A=Tháng, B-D=Doanh thu bán ra, E=Tổng DT, F=Thuế đầu ra,
    # G=DS mua vào, H=Thuế đầu vào, I=DC tăng, J=DC giảm, K=Số thuế phải nộp/khấu trừ,
    # L=blank, M=Số thuế phải nộp, N=blank, O=Thuế GTGT thuê TS, P=Ngày CK,
    # Q=Thuế TNCN thuê TS, R=Ngày CK, S=Thuế TNCN tiền lương, T=Ngày CK
    col_widths = [10, 18, 18, 18, 18, 18, 22, 18, 14, 14, 22, 4, 18, 4, 20, 16, 20, 16, 22, 16]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    def _border_cell(cell, font=None, fill=None, align=None, num_fmt=None, value=None):
        cell.border = _BORDER
        if font:    cell.font      = font
        if fill:    cell.fill      = fill
        if align:   cell.alignment = align
        if num_fmt: cell.number_format = num_fmt
        if value is not None: cell.value = value

    def _merge(r, c1, c2, value="", font=None, fill=None, align=None):
        ws.merge_cells(start_row=r, start_column=c1, end_row=r, end_column=c2)
        c = ws.cell(row=r, column=c1, value=value)
        c.border = _BORDER
        if font:  c.font      = font
        if fill:  c.fill      = fill
        if align: c.alignment = align

    # ── Rows 1-3: company info ─────────────────────────────────────────
    last = get_column_letter(len(col_widths))
    for row, text in [
        (1, f"Tên công ty: {company['name']}"),
        (2, f"Mã số thuế: {company['tax_code']}"),
        (3, f"Địa chỉ  : {company['address']}"),
    ]:
        ws.merge_cells(f"A{row}:{last}{row}")
        c = ws[f"A{row}"]
        c.value     = text
        c.font      = _BOLD10 if row == 1 else _NORM10
        c.alignment = _LEFT
        ws.row_dimensions[row].height = 15

    # Row 4: title
    ws.merge_cells(f"A4:{last}4")
    c4           = ws["A4"]
    c4.value     = "BẢNG KÊ TỔNG HỢP"
    c4.font      = _TITLE_F
    c4.alignment = _CENTER
    ws.row_dimensions[4].height = 28

    # ── Row 5: carry-forward ──────────────────────────────────────────
    ws.merge_cells("A5:J5")
    lbl = ws["A5"]
    lbl.value     = "Kỳ trước chuyển sang còn được khấu trừ:"
    lbl.font      = _BOLD10
    lbl.fill      = _CARRY_FILL
    lbl.border    = _BORDER
    lbl.alignment = _LEFT

    carry_cell = ws.cell(row=5, column=11)  # K5
    carry_cell.font      = _BOLD10
    carry_cell.fill      = _CARRY_FILL
    carry_cell.border    = _BORDER
    carry_cell.alignment = _RIGHT
    carry_cell.number_format = _NUM_FMT
    # K5 is intentionally left blank for manual entry (carry-forward balance)

    for col in range(12, len(col_widths) + 1):
        c = ws.cell(row=5, column=col)
        c.fill   = _CARRY_FILL
        c.border = _BORDER

    # Extra columns header labels on row 5 (O,Q,S = tax labels; P,R,T = date labels)
    extra = [
        (15, "Thuế GTGT thuê TS"), (16, "Ngày chuyển khoản"),
        (17, "Thuế TNCN thuê TS"), (18, "Ngày chuyển khoản"),
        (19, "Thuế TNCN tiền lương"), (20, "Ngày chuyển khoản"),
    ]
    for col, text in extra:
        c = ws.cell(row=5, column=col)
        c.value     = text
        c.font      = _BOLD10
        c.fill      = _CARRY_FILL
        c.border    = _BORDER
        c.alignment = _CENTER

    ws.row_dimensions[5].height = 32

    # ── Row 6: column headers ─────────────────────────────────────────
    headers = [
        (1,  1,  "Tháng"),
        (2,  4,  "Doanh thu bán ra 8%/10%\nDoanh thu bán ra 5%\nDoanh thu bán ra KCT"),
        (5,  5,  "Tổng Doanh Thu"),
        (6,  6,  "Thuế GTGT đầu ra"),
        (7,  7,  "Doanh số mua vào\nchưa VAT"),
        (8,  8,  "Thuế GTGT đầu vào"),
        (9,  9,  "DC TĂNG"),
        (10, 10, "DC GIẢM"),
        (11, 11, "Số thuế phải nộp\nkhấu trừ"),
        (12, 12, ""),
        (13, 13, "Số thuế\nphải nộp"),
    ]
    for c1, c2, text in headers:
        if c1 == c2:
            cell = ws.cell(row=6, column=c1, value=text)
            _border_cell(cell, font=_WHITE_BOLD, fill=_HEADER_FILL, align=_CENTER)
        else:
            _merge(6, c1, c2, text, font=_WHITE_BOLD, fill=_HEADER_FILL, align=_CENTER)
    for col in range(14, len(col_widths) + 1):
        c = ws.cell(row=6, column=col)
        _border_cell(c, font=_WHITE_BOLD, fill=_HEADER_FILL, align=_CENTER)
    ws.row_dimensions[6].height = 48

    # Sub-headers for columns 2-4
    sub_row = 7
    sub_labels = ["Doanh thu bán ra 8% 10%", "Doanh thu bán ra 5%", "Doanh thu bán ra KCT"]
    for i, label in enumerate(sub_labels, 2):
        c = ws.cell(row=sub_row, column=i, value=label)
        _border_cell(c, font=_WHITE_BOLD, fill=_HEADER_FILL, align=_CENTER)
    # Span remaining header cells
    for col in [1, 5, 6, 7, 8, 9, 10, 11, 12, 13] + list(range(14, len(col_widths)+1)):
        c = ws.cell(row=sub_row, column=col)
        _border_cell(c, font=_WHITE_BOLD, fill=_HEADER_FILL, align=_CENTER)
    ws.row_dimensions[sub_row].height = 32
    ws.freeze_panes = "A8"

    # ── Aggregate by month ────────────────────────────────────────────
    # Build per-month buckets keyed by "T{MM}" derived from issue_date
    month_order: list = []
    month_data: dict  = defaultdict(lambda: {
        "sale_taxable": 0.0, "sale_5pct": 0.0, "sale_kct": 0.0,
        "sale_vat": 0.0, "pur_amount": 0.0, "pur_vat": 0.0,
        "dc_tang": 0.0, "dc_giam": 0.0,
    })

    for inv in invoices:
        # Group by tháng của ngày lập hóa đơn (issue_date)
        if inv.issue_date:
            month = f"T{inv.issue_date.month:02d}"
        else:
            month = ""
        if month not in month_order:
            month_order.append(month)
        d     = month_data[month]
        is_sale = (inv.invoice_category or "").startswith("sale_")
        amt   = inv.amount       or 0.0
        vat   = inv.vat_amount   or 0.0
        rate  = inv.vat_rate or ""
        rate_str = str(rate).strip().upper()

        if is_sale:
            if rate_str in ("KCT", ""):
                d["sale_kct"] += amt
            elif rate_str in ("0.05", "5", "5%"):
                d["sale_5pct"] += amt
            else:
                d["sale_taxable"] += amt
            d["sale_vat"] += vat
        else:
            d["pur_amount"] += amt
            d["pur_vat"]    += vat

    # Write data rows
    data_start = 8
    current_row = data_start
    months_in_year = [f"T{i:02d}" for i in range(1, 13)]
    # Use months_in_year order, falling back to month_order for unlisted
    all_months = months_in_year + [m for m in month_order if m not in months_in_year]

    grand = defaultdict(float)

    for month_label in all_months:
        d        = month_data[month_label]
        has_data = any([
            d["sale_taxable"],
            d["sale_5pct"],
            d["sale_kct"],
            d["sale_vat"],
            d["pur_amount"],
            d["pur_vat"],
            d["dc_tang"],
            d["dc_giam"],
        ])

        if not has_data:
            continue
        
        alt_fill = _MONTH_FILL if (current_row - data_start) % 2 == 0 else None
        row      = current_row

        # Display label: strip leading "T" if it matches T01..T12, else show as-is
        display = month_label.lstrip("T").lstrip("0") if month_label.startswith("T") else month_label
        try:
            display = str(int(display))
        except Exception:
            display = month_label

        vals = {
            1:  display,
            2:  d["sale_taxable"] or None,
            3:  d["sale_5pct"]    or None,
            4:  d["sale_kct"]     or None,
            5:  f"=B{row}+C{row}+D{row}",
            6:  d["sale_vat"]     or None,
            7:  d["pur_amount"]   or None,
            8:  d["pur_vat"]      or None,
            9:  d["dc_tang"]      or None,
            10: d["dc_giam"]      or None,
            # K: carry_prev + pur_vat - sale_vat - dc_giam + dc_tang
            # Row 8 references K5; subsequent rows reference previous K
            11: (f"=K5+H{row}-F{row}-J{row}+I{row}" if row == data_start
                 else f"=K{row-1}+H{row}-F{row}-J{row}+I{row}"),
            12: None,
            13: f"=IF(K{row}<0,-K{row},0)",
        }

        for col, val in vals.items():
            c = ws.cell(row=row, column=col, value=val)
            c.border    = _BORDER
            c.font      = _NORM10
            c.alignment = _RIGHT if col != 1 else _CENTER
            if alt_fill: c.fill = alt_fill
            if col in (2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 13):
                c.number_format = _NUM_FMT

        # Extra tax payment columns (O-T) — blank, for manual entry
        for col in range(14, len(col_widths) + 1):
            c = ws.cell(row=row, column=col)
            c.border    = _BORDER
            c.font      = _NORM10
            if alt_fill: c.fill = alt_fill

        # Accumulate grand totals (numeric fields only)
        grand["sale_taxable"] += d["sale_taxable"]
        grand["sale_5pct"]    += d["sale_5pct"]
        grand["sale_kct"]     += d["sale_kct"]
        grand["sale_vat"]     += d["sale_vat"]
        grand["pur_amount"]   += d["pur_amount"]
        grand["pur_vat"]      += d["pur_vat"]
        grand["dc_tang"]      += d["dc_tang"]
        grand["dc_giam"]      += d["dc_giam"]

        ws.row_dimensions[row].height = 18
        current_row += 1

    # ── Grand total row ───────────────────────────────────────────────
    tot = current_row
    _merge(tot, 1, 1, "CỘNG", font=Font(name="Arial", bold=True, size=10),
           fill=_TOTAL_F, align=_CENTER)

    total_vals = {
        2:  grand["sale_taxable"],
        3:  grand["sale_5pct"],
        4:  grand["sale_kct"],
        5:  f"=SUM(E{data_start}:E{tot-1})",
        6:  grand["sale_vat"],
        7:  grand["pur_amount"],
        8:  grand["pur_vat"],
        9:  grand["dc_tang"],
        10: grand["dc_giam"],
        11: f"=K{tot-1}",
        12: None,
        13: f"=SUM(M{data_start}:M{tot-1})",
    }
    bold_font = Font(name="Arial", bold=True, size=10)
    for col, val in total_vals.items():
        c = ws.cell(row=tot, column=col, value=val)
        c.border         = _BORDER
        c.font           = bold_font
        c.fill           = _TOTAL_F
        c.alignment      = _RIGHT
        c.number_format  = _NUM_FMT
    for col in range(14, len(col_widths) + 1):
        c = ws.cell(row=tot, column=col)
        c.border = _BORDER
        c.fill   = _TOTAL_F
        c.font   = bold_font
    ws.row_dimensions[tot].height = 20

    # ── Note rows ─────────────────────────────────────────────────────
    note_row = tot + 2
    ws.merge_cells(f"E{note_row}:F{note_row}")
    ws[f"E{note_row}"].value     = "NOTE:"
    ws[f"E{note_row}"].font      = _BOLD10
    ws[f"E{note_row}"].alignment = _LEFT

    ws.merge_cells(f"G{note_row}:K{note_row}")
    ws[f"G{note_row}"].value     = "dương ko nộp thuế"
    ws[f"G{note_row}"].font      = _NORM10
    ws[f"G{note_row}"].alignment = _LEFT

    note_row2 = note_row + 1
    ws.merge_cells(f"G{note_row2}:K{note_row2}")
    ws[f"G{note_row2}"].value     = 'âm nộp thuế     "-"  / "()"'
    ws[f"G{note_row2}"].font      = _NORM10
    ws[f"G{note_row2}"].alignment = _LEFT

    note_row3 = note_row + 2
    ws.merge_cells(f"G{note_row3}:M{note_row3}")
    ws[f"G{note_row3}"].value     = "tồn đầu kỳ + Vat vào - Vat ra - DC giảm + DC tăng"
    ws[f"G{note_row3}"].font      = _NORM10
    ws[f"G{note_row3}"].alignment = _LEFT


def _parse_date(value: Optional[str]) -> Optional[datetime]:
    if not value:
        return None
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
        try:
            return datetime.strptime(value.strip(), fmt)
        except ValueError:
            continue
    return None